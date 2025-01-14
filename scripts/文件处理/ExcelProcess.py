import os
import requests
import xlrd
import xlwt
import logging
import time

from typing import Optional
from openpyxl import load_workbook

logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)


class ExcelProcess:
    parent_directory = os.path.dirname(os.path.abspath(__file__))

    def __init__(self, file_name: Optional[str] = None):
        """
        :param file_name: Web url or file name
        """
        self.file_name = file_name

    @classmethod
    def save_data(cls, data: list) -> None:
        # 创建一个新的工作簿
        wb = xlwt.Workbook()
        # 添加一个工作表
        ws = wb.add_sheet('Sheet1')

        # 写入表头
        ws.write(0, 0, 'col1')
        ws.write(0, 1, 'col2')
        ws.write(0, 2, 'col3')

        # 写入数据
        for idx, row in enumerate(data, start=1):
            ws.write(idx, 0, row['order_no'])
            ws.write(idx, 1, row['resp'])
            ws.write(idx, 2, row['error'])
        gitignore_out_dir = os.path.join(cls.parent_directory, 'gitignore_out')
        if not os.path.exists(gitignore_out_dir):
            os.makedirs(gitignore_out_dir)
        output_file_path = os.path.join(gitignore_out_dir, f'{time.strftime("%Y%m%d%H%M%S", time.localtime())}.xls')
        wb.save(output_file_path)

    @staticmethod
    def _open_excel_file(file_path: str):
        """
        Open the Excel file depending on its extension.
        :param file_path: The path to the Excel file.
        :return: Workbook object
        """
        # 从 xlrd 版本 2.0 起，它不再支持 .xlsx 格式
        if file_path.endswith(".xls"):
            return xlrd.open_workbook(file_path)  # For .xls files
        elif file_path.endswith(".xlsx"):
            return load_workbook(file_path)  # For .xlsx files
        else:
            raise ValueError("Unsupported file format")

    @property
    def file_path(self) -> str:
        if self.file_name is None:
            raise ValueError("File name is not set")
        gitignore_in_dir = os.path.join(self.parent_directory, 'gitignore_in')
        if self.file_name.startswith("http"):
            resp = requests.get(self.file_name)
            if resp.status_code != 200:
                raise ValueError("File download failed")
            if not os.path.exists(gitignore_in_dir):
                os.makedirs(gitignore_in_dir)
            file_path = os.path.join(gitignore_in_dir, f'{time.strftime("%Y%m%d%H%M%S", time.localtime())}.xlsx')
            with open(file_path, 'wb') as f:
                f.write(resp.content)
            logging.info('文件下载成功')
            return file_path
        return f'{gitignore_in_dir}/{self.file_name}'

    def data(self, exclude_first_row: bool = False) -> list:
        """
        按行读取所有数据
        :param exclude_first_row: 排除第一列数据
        :return:
        """
        file = self._open_excel_file(self.file_path)
        if isinstance(file, xlrd.Book):
            sheet = file.sheet_by_index(0)
            rows = sheet.nrows
            result = [sheet.row_values(row_idx) for row_idx in range(rows)]
        else:
            sheet = file.active
            rows = sheet.max_row
            result = [[sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                      for row_idx in range(1, rows + 1)]
        if exclude_first_row:
            return result[1:]
        return result

    def col_data(self, col_idx: int, exclude_first_row: bool = False) -> list:
        """
        按行读取某一列数据
        :param col_idx: 第几列
        :param exclude_first_row: 排除第一列数据
        :return:
        """
        file = self._open_excel_file(self.file_path)
        if isinstance(file, xlrd.Book):
            sheet = file.sheet_by_index(0)
            rows = sheet.nrows
            result = [sheet.cell_value(row_idx, col_idx) for row_idx in range(rows)]
        else:
            sheet = file.active
            rows = sheet.max_row
            result = [sheet.cell(row=row_idx, column=col_idx + 1).value for row_idx in range(1, rows + 1)]
        if exclude_first_row:
            return result[1:]
        return result
